import openpyxl
import os 

directory = '..../excel/'

wb= openpyxl.load_workbook(directory+"octobre.xlsx", data_only=True)
wb1= openpyxl.load_workbook(directory+"novembre.xlsx", data_only=True)
wb2= openpyxl.load_workbook(directory+"decembre.xlsx", data_only=True)

def add_data_from_wb(wb,d):

    sheet = wb.active
    donnees = {}
    for row in range(2, sheet.max_row):
        #print(sheet1.cell(row, 1).value)
        nom_article = sheet.cell(row,1).value
        if not nom_article:
            break
        total_ventes = sheet.cell(row, 4).value
        if d.get(nom_article):
            d[nom_article].append(total_ventes)
        else:
            d[nom_article] = [total_ventes]

donnees = {}
add_data_from_wb(wb, donnees)
add_data_from_wb(wb1, donnees)
add_data_from_wb(wb2, donnees)

print(donnees)

wb_sorti= openpyxl.Workbook()
sheet = wb_sorti.active
sheet["A1"] ="Article"
sheet["B1"] ="Octobre"
sheet["C1"] ="Novembre"
sheet["D1"] ="Decembre"

row= 2
for i in donnees.items():
    print(i)
    nom_article = i[0]
    ventes = i[1]
    sheet.cell(row, 1).value = nom_article
    for j in range(0, len(ventes)):
        sheet.cell(row, 2+j).value = ventes[j]

    row +=1

wb_sorti.save(directory+"total_vente_trimestre.xlsx")

chart_ref = openpyxl.chart.Reference(sheet, min_col = 2 ,min_row = 2 , max_col = sheet.max_column ,max_row = 2)
chart_serie = openpyxl.chart.Series(chart_ref, title="Total ventes en €")
chart = openpyxl.chart.BarChart()
chart.title = "Evolution du prix des pommes"
chart.append(chart_serie)
sheet.add_chart(chart, "F2")

wb_sorti.save(directory+"total_vente_trimestre.xlsx")
